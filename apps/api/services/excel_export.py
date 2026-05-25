"""Excel quantity table export via openpyxl."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER = ["Year", "Pit area (m²)", "Excavation (m³)", "Mineral (m³)",
          "Mineral (t)", "Saleable (t)", "Topsoil (m³)", "Overburden (m³)",
          "Backfill (m³)", "Plantation (m²)", "Stripping ratio"]
KEYS = ["year", "pit_area_m2", "excavation_volume_m3", "mineral_volume_m3",
        "mineral_tonnes", "saleable_tonnes", "topsoil_m3", "overburden_m3",
        "backfill_m3", "plantation_area_m2", "stripping_ratio"]


def project_to_excel(project: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = "Summary"

    details = project.get("project_details") or {}
    ws.append(["Mining Plan Quantity Tables"])
    ws["A1"].font = Font(bold=True, size=14, color="0F172A")
    ws.append([])
    for k in ("project_name", "applicant_name", "mineral", "village", "tehsil",
              "district", "state", "area_ha", "scale", "plan_period_years"):
        ws.append([k.replace("_", " ").title(), details.get(k, "")])
    ws.append([])

    for qt in project.get("quantity_tables") or []:
        title = f"Quantity Table — {qt.get('alternative', 'plan')}"
        sheet_name = (qt.get("alternative") or "plan")[:31]
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        s = wb.create_sheet(sheet_name)
        s.append([title])
        s["A1"].font = Font(bold=True, size=12)
        s.append([])
        s.append(HEADER)
        for col_i, _ in enumerate(HEADER, start=1):
            cell = s.cell(row=3, column=col_i)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="0F172A")
            cell.alignment = Alignment(horizontal="center")
        for row in qt.get("rows") or []:
            s.append([row.get(k) for k in KEYS])
        for col_i, h in enumerate(HEADER, start=1):
            s.column_dimensions[get_column_letter(col_i)].width = max(12, len(h) + 2)

    wv = project.get("validation_warnings") or []
    if wv:
        s = wb.create_sheet("Validation")
        s.append(["Severity", "Code", "Alternative", "Message"])
        for col_i in range(1, 5):
            c = s.cell(row=1, column=col_i)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="0F172A")
        for w in wv:
            s.append([w.get("severity", ""), w.get("code", ""), w.get("alternative", ""), w.get("message", "")])
        for col_i, w in enumerate([12, 18, 18, 80], start=1):
            s.column_dimensions[get_column_letter(col_i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
